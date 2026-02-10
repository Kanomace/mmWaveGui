import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Base folder path
base_folder_path = r'C:\Users\Kano\Desktop\radar_toolbox_1_30_01_03\tools\visualizers\Industrial_Visualizer\binData\traindata\9sign'

# Number of subfolders
folder_count = 50

# Create a new Excel workbook
combined_workbook = Workbook()
combined_sheet = combined_workbook.active

# Iterate over subfolders
for i in range(1, folder_count + 1):
    folder_path = os.path.join(base_folder_path, str(i))
    print(f'Reading Excel files from folder {i}')

    # Create an empty list to store data from each subfolder
    data_lists = []

    # Iterate over subfolder suffixes
    for suffix in ['pHistBytes_clustered_voxel_XOZ', 'pHistBytes_clustered_voxel_YOZ']:
        # Full subfolder path
        full_folder_path = os.path.join(folder_path, suffix)

        # Check whether the Excel file exists in the subfolder
        excel_file_path = os.path.join(full_folder_path, 'image_data.xlsx')
        if not os.path.exists(excel_file_path):
            continue

        # Open the Excel file
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # Read data from the Excel file
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Append the data to the list
        data_lists.append(data)

    # Check whether data from both subfolders exists
    if len(data_lists) == 2:
        # Write header row
        if i == 1:
            combined_sheet.append(data_lists[0][0] + data_lists[1][0])

        # Concatenate data into the new Excel workbook (excluding headers)
        for j in range(1, len(data_lists[0])):
            combined_sheet.append(data_lists[0][j] + data_lists[1][j])

# Adjust column widths
for column in combined_sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.2
    combined_sheet.column_dimensions[column_letter].width = adjusted_width

# Save the merged Excel file
combined_file_path = os.path.join(base_folder_path, 'combined_image_data.xlsx')
combined_workbook.save(combined_file_path)

print('The merged Excel file has been saved to:', combined_file_path)
